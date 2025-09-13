# export_utils.py - Utilitaires d'export PDF/Excel pour AppGET
import io
import os
from datetime import datetime, date, timedelta
from typing import List, Tuple
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.conf import settings


def export_schedule_to_pdf(schedules, title: str, start_date: date, end_date: date) -> Tuple[bytes, str]:
    """Exporter un emploi du temps en PDF"""
    
    # Créer un buffer pour le PDF
    buffer = io.BytesIO()
    
    # Configuration de la page
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=1*inch,
        bottomMargin=0.5*inch
    )
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.darkgreen
    )
    
    # Contenu du document
    story = []
    
    # Titre principal
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(
        f"Semaine du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}", 
        subtitle_style
    ))
    story.append(Spacer(1, 20))
    
    # Organiser les données par jour
    days_data = {}
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    # Initialiser les jours
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        days_data[i] = {
            'name': day_names[i],
            'date': current_date,
            'schedules': []
        }
    
    # Remplir les données
    for schedule in schedules:
        day = schedule.time_slot.day_of_week
        if day in days_data:
            days_data[day]['schedules'].append(schedule)
    
    # Créer les tableaux par jour
    for day_num, day_info in days_data.items():
        if not day_info['schedules']:
            continue
            
        # Titre du jour
        day_title = f"{day_info['name']} {day_info['date'].strftime('%d/%m/%Y')}"
        story.append(Paragraph(day_title, styles['Heading3']))
        story.append(Spacer(1, 10))
        
        # Trier par heure de début
        day_schedules = sorted(day_info['schedules'], key=lambda x: x.time_slot.start_time)
        
        # Créer le tableau
        table_data = [
            ['Heure', 'Matière', 'Enseignant', 'Salle', 'Programmes', 'Type']
        ]
        
        for schedule in day_schedules:
            programs_names = ', '.join([p.name for p in schedule.programs.all()])
            
            table_data.append([
                f"{schedule.time_slot.start_time.strftime('%H:%M')}-{schedule.time_slot.end_time.strftime('%H:%M')}",
                schedule.subject.name,
                schedule.teacher.user.get_full_name(),
                schedule.room.name,
                programs_names[:30] + '...' if len(programs_names) > 30 else programs_names,
                schedule.subject.get_subject_type_display()
            ])
        
        # Style du tableau
        table = Table(table_data, colWidths=[0.8*inch, 1.8*inch, 1.5*inch, 1*inch, 1.5*inch, 0.8*inch])
        table.setStyle(TableStyle([
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Corps du tableau
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
            
            # Bordures
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alternance de couleurs pour les lignes
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white]),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    # Ajouter des statistiques
    story.append(PageBreak())
    story.append(Paragraph("Statistiques", styles['Heading2']))
    story.append(Spacer(1, 10))
    
    total_sessions = len(schedules)
    total_hours = sum([schedule.duration_minutes for schedule in schedules]) / 60
    unique_subjects = len(set([schedule.subject.id for schedule in schedules]))
    unique_teachers = len(set([schedule.teacher.id for schedule in schedules]))
    unique_rooms = len(set([schedule.room.id for schedule in schedules]))
    
    stats_data = [
        ['Statistique', 'Valeur'],
        ['Nombre total de séances', str(total_sessions)],
        ['Total d\'heures', f"{total_hours:.1f}h"],
        ['Matières différentes', str(unique_subjects)],
        ['Enseignants impliqués', str(unique_teachers)],
        ['Salles utilisées', str(unique_rooms)],
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 1.5*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgreen, colors.white]),
    ]))
    
    story.append(stats_table)
    
    # Pied de page
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - AppGET",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
    ))
    
    # Générer le PDF
    doc.build(story)
    
    # Récupérer le contenu
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Nom du fichier
    filename = f"emploi_du_temps_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
    
    return pdf_data, filename


def export_schedule_to_excel(schedules, title: str, start_date: date, end_date: date) -> Tuple[bytes, str]:
    """Exporter un emploi du temps en Excel"""
    
    # Créer un buffer pour le fichier Excel
    buffer = io.BytesIO()
    
    # Organiser les données
    schedule_data = []
    
    for schedule in schedules:
        programs_names = ', '.join([p.name for p in schedule.programs.all()])
        
        schedule_data.append({
            'Date': schedule.start_date,
            'Jour': schedule.time_slot.get_day_of_week_display(),
            'Heure début': schedule.time_slot.start_time,
            'Heure fin': schedule.time_slot.end_time,
            'Durée (min)': schedule.duration_minutes,
            'Matière': schedule.subject.name,
            'Code matière': schedule.subject.code,
            'Type': schedule.subject.get_subject_type_display(),
            'Enseignant': schedule.teacher.user.get_full_name(),
            'Email enseignant': schedule.teacher.user.email,
            'Salle': schedule.room.name,
            'Type salle': schedule.room.get_room_type_display(),
            'Capacité salle': schedule.room.capacity,
            'Programmes': programs_names,
            'Nombre étudiants': schedule.student_count,
            'Statut': 'Actif' if schedule.is_active else 'Inactif',
            'Annulé': 'Oui' if schedule.is_cancelled else 'Non',
            'Rattrapage': 'Oui' if schedule.is_makeup else 'Non',
            'Notes': schedule.notes or ''
        })
    
    # Créer le DataFrame
    df = pd.DataFrame(schedule_data)
    
    # Créer un writer Excel avec plusieurs feuilles
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        
        # Feuille principale avec tous les emplois du temps
        if not df.empty:
            df.to_excel(writer, sheet_name='Emploi du temps', index=False)
            
            # Formatter la feuille principale
            worksheet = writer.sheets['Emploi du temps']
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Style de l'en-tête
            from openpyxl.styles import PatternFill, Font, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        # Feuille avec les statistiques
        stats_data = {
            'Statistique': [
                'Période',
                'Nombre total de séances',
                'Total d\'heures',
                'Matières différentes',
                'Enseignants impliqués',
                'Salles utilisées',
                'Programmes concernés'
            ],
            'Valeur': [
                f"Du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}",
                len(schedules),
                f"{sum([schedule.duration_minutes for schedule in schedules]) / 60:.1f}h",
                len(set([schedule.subject.id for schedule in schedules])),
                len(set([schedule.teacher.id for schedule in schedules])),
                len(set([schedule.room.id for schedule in schedules])),
                len(set([prog.id for schedule in schedules for prog in schedule.programs.all()]))
            ]
        }
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistiques', index=False)
        
        # Feuille avec le planning par jour
        if not df.empty:
            # Créer une vue par jour
            days_data = {}
            
            for _, row in df.iterrows():
                day = row['Jour']
                if day not in days_data:
                    days_data[day] = []
                days_data[day].append(row)
            
            # Créer une feuille pour chaque jour avec des données
            for day, day_schedules in days_data.items():
                if day_schedules:
                    day_df = pd.DataFrame(day_schedules)
                    # Trier par heure
                    day_df = day_df.sort_values('Heure début')
                    
                    sheet_name = f"{day}"[:31]  # Excel limite à 31 caractères
                    day_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Formatter la feuille du jour
                    day_worksheet = writer.sheets[sheet_name]
                    
                    for cell in day_worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
        
        # Feuille avec les conflits potentiels (si il y en a)
        conflicts = []
        
        # Détecter les conflits de salles
        for schedule1 in schedules:
            for schedule2 in schedules:
                if (schedule1.id != schedule2.id and 
                    schedule1.room.id == schedule2.room.id and
                    schedule1.time_slot.id == schedule2.time_slot.id and
                    schedule1.start_date == schedule2.start_date):
                    
                    conflicts.append({
                        'Type conflit': 'Salle',
                        'Ressource': schedule1.room.name,
                        'Date': schedule1.start_date,
                        'Créneau': str(schedule1.time_slot),
                        'Cours 1': f"{schedule1.subject.name} - {schedule1.teacher.user.get_full_name()}",
                        'Cours 2': f"{schedule2.subject.name} - {schedule2.teacher.user.get_full_name()}",
                    })
        
        # Détecter les conflits d'enseignants
        for schedule1 in schedules:
            for schedule2 in schedules:
                if (schedule1.id != schedule2.id and 
                    schedule1.teacher.id == schedule2.teacher.id and
                    schedule1.time_slot.id == schedule2.time_slot.id and
                    schedule1.start_date == schedule2.start_date):
                    
                    conflicts.append({
                        'Type conflit': 'Enseignant',
                        'Ressource': schedule1.teacher.user.get_full_name(),
                        'Date': schedule1.start_date,
                        'Créneau': str(schedule1.time_slot),
                        'Cours 1': f"{schedule1.subject.name} - {schedule1.room.name}",
                        'Cours 2': f"{schedule2.subject.name} - {schedule2.room.name}",
                    })
        
        if conflicts:
            conflicts_df = pd.DataFrame(conflicts)
            conflicts_df.to_excel(writer, sheet_name='Conflits détectés', index=False)
            
            # Formatter la feuille des conflits
            conflicts_worksheet = writer.sheets['Conflits détectés']
            
            # Colorer en rouge l'en-tête des conflits
            conflict_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            for cell in conflicts_worksheet[1]:
                cell.fill = conflict_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
    
    # Récupérer le contenu
    excel_data = buffer.getvalue()
    buffer.close()
    
    # Nom du fichier
    filename = f"emploi_du_temps_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return excel_data, filename


def create_weekly_schedule_grid(schedules, start_date: date, end_date: date) -> pd.DataFrame:
    """Créer une grille d'emploi du temps hebdomadaire"""
    
    # Créer une structure de grille
    time_slots = set()
    for schedule in schedules:
        time_slot_str = f"{schedule.time_slot.start_time.strftime('%H:%M')}-{schedule.time_slot.end_time.strftime('%H:%M')}"
        time_slots.add(time_slot_str)
    
    time_slots = sorted(list(time_slots))
    days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    # Créer le DataFrame
    grid_data = {}
    for day in days:
        grid_data[day] = [''] * len(time_slots)
    
    # Remplir la grille
    for schedule in schedules:
        day_name = schedule.time_slot.get_day_of_week_display()
        time_slot_str = f"{schedule.time_slot.start_time.strftime('%H:%M')}-{schedule.time_slot.end_time.strftime('%H:%M')}"
        
        if day_name in days and time_slot_str in time_slots:
            time_index = time_slots.index(time_slot_str)
            course_info = f"{schedule.subject.name}\n{schedule.teacher.user.get_full_name()}\n{schedule.room.name}"
            grid_data[day_name][time_index] = course_info
    
    # Créer le DataFrame final
    df = pd.DataFrame(grid_data, index=time_slots)
    
    return df


def export_weekly_grid_to_excel(schedules, title: str, start_date: date, end_date: date) -> Tuple[bytes, str]:
    """Exporter un emploi du temps sous forme de grille hebdomadaire"""
    
    buffer = io.BytesIO()
    
    # Créer la grille
    grid_df = create_weekly_schedule_grid(schedules, start_date, end_date)
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        
        # Écrire la grille
        grid_df.to_excel(writer, sheet_name='Grille hebdomadaire', index=True)
        
        # Formatter la feuille
        worksheet = writer.sheets['Grille hebdomadaire']
        
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Style de l'en-tête
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # En-têtes des colonnes (jours)
        for col in range(2, len(grid_df.columns) + 2):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # En-têtes des lignes (créneaux horaires)
        for row in range(2, len(grid_df.index) + 2):
            cell = worksheet.cell(row=row, column=1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Formater les cellules de contenu
        for row in range(2, len(grid_df.index) + 2):
            for col in range(2, len(grid_df.columns) + 2):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Bordures
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
                
                # Couleur de fond alternée
                if (row + col) % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # Ajuster les dimensions
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 5, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Hauteur des lignes
        for row_num in range(2, len(grid_df.index) + 2):
            worksheet.row_dimensions[row_num].height = 60
    
    excel_data = buffer.getvalue()
    buffer.close()
    
    filename = f"grille_emploi_temps_{start_date.strftime('%Y%m%d')}.xlsx"
    
    return excel_data, filename
