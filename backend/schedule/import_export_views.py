from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from django.utils import timezone
from django.core.exceptions import ValidationError
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import io
import csv
from calendar import Calendar

from .models import Schedule
from core.models import Department, Program, Room, Subject, Teacher
from authentication.models import User


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_schedules(request):
    """
    Importe des emplois du temps à partir d'un fichier Excel/CSV
    """
    if 'file' not in request.FILES:
        return Response(
            {'success': False, 'message': 'Aucun fichier fourni'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    file = request.FILES['file']
    
    # Vérifier l'extension du fichier
    allowed_extensions = ['.xlsx', '.xls', '.csv']
    file_extension = '.' + file.name.split('.')[-1].lower()
    
    if file_extension not in allowed_extensions:
        return Response({
            'success': False,
            'message': f'Format de fichier non supporté: {file_extension}',
            'errors': [f'Utilisez un fichier Excel (.xlsx, .xls) ou CSV (.csv)']
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Lire le fichier selon son type
        if file_extension == '.csv':
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # Vérifier les colonnes requises
        required_columns = [
            'titre', 'matiere', 'enseignant', 'salle', 'programme',
            'jour_semaine', 'heure_debut', 'heure_fin', 'date_debut', 'date_fin'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return Response({
                'success': False,
                'message': 'Colonnes manquantes dans le fichier',
                'errors': [f'Colonnes manquantes: {", ".join(missing_columns)}']
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Traiter les données
        imported_count = 0
        errors = []
        warnings = []
        
        for index, row in df.iterrows():
            try:
                # Valider et récupérer les objets liés
                try:
                    subject = Subject.objects.get(name__iexact=str(row['matiere']).strip())
                except Subject.DoesNotExist:
                    errors.append(f"Ligne {index + 2}: Matière '{row['matiere']}' non trouvée")
                    continue
                
                try:
                    teacher = Teacher.objects.get(user__full_name__icontains=str(row['enseignant']).strip())
                except Teacher.DoesNotExist:
                    errors.append(f"Ligne {index + 2}: Enseignant '{row['enseignant']}' non trouvé")
                    continue
                
                try:
                    room = Room.objects.get(name__iexact=str(row['salle']).strip())
                except Room.DoesNotExist:
                    errors.append(f"Ligne {index + 2}: Salle '{row['salle']}' non trouvée")
                    continue
                
                try:
                    program = Program.objects.get(name__icontains=str(row['programme']).strip())
                except Program.DoesNotExist:
                    errors.append(f"Ligne {index + 2}: Programme '{row['programme']}' non trouvé")
                    continue
                
                # Convertir le jour de la semaine
                day_mapping = {
                    'lundi': 0, 'mardi': 1, 'mercredi': 2, 'jeudi': 3, 'vendredi': 4, 'samedi': 5
                }
                day_str = str(row['jour_semaine']).lower().strip()
                if day_str not in day_mapping:
                    errors.append(f"Ligne {index + 2}: Jour '{row['jour_semaine']}' invalide")
                    continue
                
                day_of_week = day_mapping[day_str]
                
                # Convertir les heures
                try:
                    start_time = pd.to_datetime(str(row['heure_debut'])).time()
                    end_time = pd.to_datetime(str(row['heure_fin'])).time()
                except:
                    errors.append(f"Ligne {index + 2}: Format d'heure invalide")
                    continue
                
                # Convertir les dates
                try:
                    week_start = pd.to_datetime(row['date_debut']).date()
                    week_end = pd.to_datetime(row['date_fin']).date()
                except:
                    errors.append(f"Ligne {index + 2}: Format de date invalide")
                    continue
                
                # Vérifier les conflits avant de créer
                from django.db.models import Q
                existing_conflicts = Schedule.objects.filter(
                    Q(room=room) | Q(teacher=teacher),
                    day_of_week=day_of_week,
                    week_start=week_start,
                    start_time__lt=end_time,
                    end_time__gt=start_time,
                    is_active=True
                ).exists()
                
                if existing_conflicts:
                    warnings.append(f"Ligne {index + 2}: Conflit détecté mais cours importé")
                
                # Créer le schedule
                schedule = Schedule.objects.create(
                    title=str(row['titre']).strip(),
                    subject=subject,
                    teacher=teacher,
                    room=room,
                    program=program,
                    day_of_week=day_of_week,
                    start_time=start_time,
                    end_time=end_time,
                    week_start=week_start,
                    week_end=week_end,
                    created_by=request.user,
                    is_active=True
                )
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Ligne {index + 2}: Erreur - {str(e)}")
                continue
        
        # Préparer la réponse
        if imported_count > 0:
            return Response({
                'success': True,
                'message': f'{imported_count} cours importés avec succès',
                'imported_count': imported_count,
                'warnings': warnings[:10],  # Limiter à 10 avertissements
                'errors': errors[:10]  # Limiter à 10 erreurs
            })
        else:
            return Response({
                'success': False,
                'message': 'Aucun cours n\'a pu être importé',
                'errors': errors[:10]
            }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Erreur lors de la lecture du fichier: {str(e)}',
            'errors': [str(e)]
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_import_template(request):
    """
    Télécharge un modèle Excel pour l'import
    """
    # Créer un workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle Import"
    
    # En-têtes
    headers = [
        'titre', 'matiere', 'enseignant', 'salle', 'programme',
        'jour_semaine', 'heure_debut', 'heure_fin', 'date_debut', 'date_fin'
    ]
    
    # Descriptions des colonnes
    descriptions = [
        'Titre du cours (ex: Cours de Mathématiques)',
        'Nom exact de la matière',
        'Nom de l\'enseignant',
        'Code de la salle (ex: AMPH-A)',
        'Nom du programme (ex: L3 Informatique)',
        'Jour (lundi, mardi, mercredi, jeudi, vendredi, samedi)',
        'Heure de début (format HH:MM, ex: 08:00)',
        'Heure de fin (format HH:MM, ex: 10:00)',
        'Date de début de période (format YYYY-MM-DD)',
        'Date de fin de période (format YYYY-MM-DD)'
    ]
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Écrire les descriptions
    for col, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=col, value=desc)
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True)
    
    # Ajouter un exemple
    example_data = [
        'Programmation Orientée Objet',
        'POO',
        'Prof. Alami',
        'AMPH-A',
        'L3 Informatique',
        'lundi',
        '08:00',
        '10:00',
        '2024-09-01',
        '2024-12-31'
    ]
    
    for col, data in enumerate(example_data, 1):
        cell = ws.cell(row=3, column=col, value=data)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="modele_emploi_temps.xlsx"'
    
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_schedules(request):
    """
    Exporte les emplois du temps selon le format demandé
    """
    export_format = request.GET.get('format', 'excel')
    period = request.GET.get('period', 'current_month')
    include_details = request.GET.get('include_details', 'true').lower() == 'true'
    departments = request.GET.get('departments', '').split(',') if request.GET.get('departments') else []
    programs = request.GET.get('programs', '').split(',') if request.GET.get('programs') else []
    
    # Calculer la période
    today = timezone.now().date()
    if period == 'current_week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif period == 'current_month':
        start_date = today.replace(day=1)
        next_month = start_date.replace(month=start_date.month + 1) if start_date.month < 12 else start_date.replace(year=start_date.year + 1, month=1)
        end_date = next_month - timedelta(days=1)
    elif period == 'current_semester':
        # Supposer que le semestre commence en septembre ou février
        if today.month >= 9:
            start_date = today.replace(month=9, day=1)
            end_date = today.replace(month=12, day=31)
        elif today.month <= 6:
            start_date = today.replace(month=2, day=1)
            end_date = today.replace(month=6, day=30)
        else:
            start_date = today.replace(month=2, day=1)
            end_date = today.replace(month=6, day=30)
    else:  # current_year
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    
    # Filtrer les schedules
    schedules = Schedule.objects.filter(
        week_start__lte=end_date,
        week_end__gte=start_date,
        is_active=True
    ).select_related('subject', 'teacher__user', 'room', 'program')
    
    if departments:
        schedules = schedules.filter(subject__department__id__in=departments)
    if programs:
        schedules = schedules.filter(program__id__in=programs)
    
    if export_format == 'excel':
        return export_excel(schedules, include_details, start_date, end_date)
    elif export_format == 'pdf':
        return export_pdf(schedules, include_details, start_date, end_date)
    elif export_format == 'ics':
        return export_ics(schedules, start_date, end_date)
    else:
        return Response({'error': 'Format non supporté'}, status=status.HTTP_400_BAD_REQUEST)


def export_excel(schedules, include_details, start_date, end_date):
    """Exporte les emplois du temps en format Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emplois du Temps"
    
    # En-têtes
    if include_details:
        headers = [
            'Titre', 'Matière', 'Enseignant', 'Salle', 'Programme',
            'Jour', 'Heure Début', 'Heure Fin', 'Date Début', 'Date Fin', 'Statut'
        ]
    else:
        headers = ['Titre', 'Matière', 'Enseignant', 'Salle', 'Jour', 'Heure Début', 'Heure Fin']
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Écrire les données
    for row_num, schedule in enumerate(schedules.order_by('day_of_week', 'start_time'), 2):
        day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
        day_name = day_names[schedule.day_of_week] if schedule.day_of_week < len(day_names) else f'Jour {schedule.day_of_week}'
        
        if include_details:
            data = [
                schedule.title,
                schedule.subject.name,
                schedule.teacher.user.full_name,
                schedule.room.name,
                schedule.program.name,
                day_name,
                schedule.start_time.strftime('%H:%M'),
                schedule.end_time.strftime('%H:%M'),
                schedule.week_start.strftime('%Y-%m-%d'),
                schedule.week_end.strftime('%Y-%m-%d'),
                'Actif' if schedule.is_active else 'Inactif'
            ]
        else:
            data = [
                schedule.title,
                schedule.subject.name,
                schedule.teacher.user.full_name,
                schedule.room.name,
                day_name,
                schedule.start_time.strftime('%H:%M'),
                schedule.end_time.strftime('%H:%M')
            ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Ajuster les largeurs de colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="emplois_temps_{start_date}_{end_date}.xlsx"'
    
    return response


def export_pdf(schedules, include_details, start_date, end_date):
    """Exporte les emplois du temps en format PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title = Paragraph(f"Emplois du Temps - {start_date} à {end_date}", styles['Title'])
    elements.append(title)
    
    # Préparer les données pour le tableau
    data = [['Titre', 'Matière', 'Enseignant', 'Salle', 'Jour', 'Horaires']]
    
    for schedule in schedules.order_by('day_of_week', 'start_time'):
        day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
        day_name = day_names[schedule.day_of_week] if schedule.day_of_week < len(day_names) else f'Jour {schedule.day_of_week}'
        
        data.append([
            schedule.title[:20],  # Limiter la longueur
            schedule.subject.name[:15],
            schedule.teacher.user.full_name[:20],
            schedule.room.name,
            day_name,
            f"{schedule.start_time.strftime('%H:%M')}-{schedule.end_time.strftime('%H:%M')}"
        ])
    
    # Créer le tableau
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="emplois_temps_{start_date}_{end_date}.pdf"'
    
    return response


def export_ics(schedules, start_date, end_date):
    """Exporte les emplois du temps en format ICS (calendrier)"""
    from datetime import datetime
    import uuid
    
    # Créer le contenu ICS
    ics_content = []
    ics_content.append("BEGIN:VCALENDAR")
    ics_content.append("VERSION:2.0")
    ics_content.append("PRODID:-//AppGET//Emplois du Temps//FR")
    ics_content.append("CALSCALE:GREGORIAN")
    ics_content.append("METHOD:PUBLISH")
    
    for schedule in schedules:
        # Calculer la date complète pour chaque occurrence
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() == schedule.day_of_week:
                # Créer l'événement
                event_start = datetime.combine(current_date, schedule.start_time)
                event_end = datetime.combine(current_date, schedule.end_time)
                
                ics_content.append("BEGIN:VEVENT")
                ics_content.append(f"UID:{uuid.uuid4()}@appget.university.edu")
                ics_content.append(f"DTSTART:{event_start.strftime('%Y%m%dT%H%M%S')}")
                ics_content.append(f"DTEND:{event_end.strftime('%Y%m%dT%H%M%S')}")
                ics_content.append(f"SUMMARY:{schedule.title}")
                ics_content.append(f"DESCRIPTION:Matière: {schedule.subject.name}\\nEnseignant: {schedule.teacher.user.full_name}\\nProgramme: {schedule.program.name}")
                ics_content.append(f"LOCATION:{schedule.room.name}")
                ics_content.append(f"CREATED:{timezone.now().strftime('%Y%m%dT%H%M%SZ')}")
                ics_content.append("END:VEVENT")
            
            current_date += timedelta(days=1)
    
    ics_content.append("END:VCALENDAR")
    
    response = HttpResponse('\n'.join(ics_content), content_type='text/calendar')
    response['Content-Disposition'] = f'attachment; filename="emplois_temps_{start_date}_{end_date}.ics"'
    
    return response
