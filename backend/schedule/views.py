from rest_framework import generics
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from datetime import datetime, timedelta

from .models import Schedule, Absence, MakeupSession
from .serializers import (
    ScheduleSerializer,
    AbsenceSerializer,
    MakeupSessionSerializer,
)
from .utils import check_conflicts, get_available_rooms


class ScheduleListCreateView(generics.ListCreateAPIView):
    serializer_class = ScheduleSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['program', 'teacher', 'room', 'subject', 'day_of_week', 'is_active']
    
    def get_queryset(self):
        user = self.request.user
        queryset = Schedule.objects.all()
        
        if user.role == 'admin':
            return queryset
        elif user.role == 'department_head':
            return queryset.filter(program__department=user.department)
        elif user.role == 'program_head':
            return queryset.filter(program=user.program)
        elif user.role == 'teacher':
            return queryset.filter(teacher__user=user)
        elif user.role == 'student':
            return queryset.filter(program=user.student.program)
        
        return queryset.none()
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class ScheduleDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ScheduleSerializer
    
    def get_queryset(self):
        user = self.request.user
        queryset = Schedule.objects.all()
        
        if user.role == 'admin':
            return queryset
        elif user.role == 'department_head':
            return queryset.filter(program__department=user.department)
        elif user.role == 'program_head':
            return queryset.filter(program=user.program)
        elif user.role == 'teacher':
            return queryset.filter(teacher__user=user)
        
        return queryset.none()

@api_view(['POST'])
def check_schedule_conflicts(request):
    """Check for conflicts before creating/updating a schedule"""
    data = request.data
    conflicts = check_conflicts(
        teacher_id=data.get('teacher'),
        room_id=data.get('room'),
        day_of_week=data.get('day_of_week'),
        start_time=data.get('start_time'),
        end_time=data.get('end_time'),
        week_start=data.get('week_start'),
        week_end=data.get('week_end'),
        exclude_id=data.get('schedule_id')
    )
    
    return Response({
        'has_conflicts': len(conflicts) > 0,
        'conflicts': conflicts
    })

@api_view(['GET'])
def get_schedule_by_week(request):
    """Get schedule for a specific week formatted for frontend - CORRIGÉ"""
    # Support both week_start and start_date parameters for compatibility
    week_start = request.GET.get('week_start') or request.GET.get('start_date')
    program_id = request.GET.get('program_id') or request.GET.get('program')
    teacher_id = request.GET.get('teacher_id') or request.GET.get('teacher')
    room_id = request.GET.get('room_id') or request.GET.get('room')
    subject_id = request.GET.get('subject_id') or request.GET.get('subject')
    export_format = request.GET.get('format', 'json')
    
    if not week_start:
        return Response({'error': 'week_start or start_date parameter required'}, status=400)
    
    try:
        week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + timedelta(days=6)
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    
    # Base queryset - MODIFIÉ pour être plus flexible avec les dates
    queryset = Schedule.objects.filter(
        Q(week_start=week_start_date) | 
        Q(week_start__lte=week_start_date, week_end__gte=week_start_date),
        is_active=True
    ).select_related('subject', 'teacher__user', 'room', 'program')
    
    # Apply filters based on parameters
    if program_id:
        queryset = queryset.filter(program_id=program_id)
    if teacher_id:
        queryset = queryset.filter(teacher_id=teacher_id)
    if room_id:
        queryset = queryset.filter(room_id=room_id)
    if subject_id:
        queryset = queryset.filter(subject_id=subject_id)
    
    # Apply user role permissions
    user = request.user
    if user.role == 'department_head':
        queryset = queryset.filter(program__department=user.department)
    elif user.role == 'program_head':
        queryset = queryset.filter(program=user.program)
    elif user.role == 'teacher':
        queryset = queryset.filter(teacher__user=user)
    elif user.role == 'student':
        try:
            queryset = queryset.filter(program=user.student.program)
        except Exception:
            queryset = queryset.none()
    
    # Organize data by days
    days_data = []
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    total_sessions = 0
    total_minutes = 0
    
    for day_index in range(7):
        current_date = week_start_date + timedelta(days=day_index)
        day_schedules = []
        
        # Get schedules for this day
        for schedule in queryset.filter(day_of_week=day_index):
            total_sessions += 1
            duration_minutes = (
                datetime.combine(datetime.today(), schedule.end_time)
                - datetime.combine(datetime.today(), schedule.start_time)
            ).seconds // 60
            total_minutes += duration_minutes
            
            # Create schedule data structure matching frontend expectations
            schedule_data = {
                'id': schedule.id,
                'title': schedule.title,
                'subject_name': schedule.subject.name,
                'subject_code': getattr(schedule.subject, 'code', ''),
                'teacher_name': schedule.teacher.user.get_full_name(),
                'room_name': schedule.room.name,
                'room_capacity': schedule.room.capacity,
                'time_slot_info': {
                    'id': schedule.id,
                    'day_of_week': schedule.day_of_week,
                    'day_display': schedule.get_day_of_week_display(),
                    'start_time': schedule.start_time.strftime('%H:%M'),
                    'end_time': schedule.end_time.strftime('%H:%M'),
                    'duration_minutes': duration_minutes
                },
                'programs_list': [schedule.program.name],
                'start_date': current_date.strftime('%Y-%m-%d'),
                'end_date': current_date.strftime('%Y-%m-%d'),
                'duration_minutes': duration_minutes,
                'student_count': getattr(schedule.program, 'capacity', 0),
                'is_room_suitable': True,
                'is_cancelled': not schedule.is_active,
                'is_makeup': False,
                'notes': schedule.notes
            }
            day_schedules.append(schedule_data)
        
        # Sort schedules by start time
        day_schedules.sort(key=lambda x: x['time_slot_info']['start_time'])
        
        days_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'day_name': day_names[day_index],
            'schedules': day_schedules
        })
    
    # Calculate total hours
    total_hours = total_minutes / 60.0 if total_minutes > 0 else 0
    
    # Response data structure matching frontend expectations
    response_data = {
        'week_start': week_start_date.strftime('%Y-%m-%d'),
        'week_end': week_end_date.strftime('%Y-%m-%d'),
        'days': days_data,
        'total_sessions': total_sessions,
        'total_hours': total_hours
    }
    
    # Handle different export formats
    if export_format == 'pdf':
        return export_week_pdf(queryset, week_start_date, week_end_date)
    elif export_format == 'excel':
        return export_week_excel(queryset, week_start_date, week_end_date)
    else:
        return Response(response_data)


@api_view(['GET'])
def get_schedule_by_range(request):
    """Get schedules for a date range [start_date, end_date] inclusive.

    Query params:
      - start_date (YYYY-MM-DD) required
      - end_date (YYYY-MM-DD) required
      - program_id, teacher_id, room_id, subject_id optional
      - format: json|pdf|excel (default json)
    """
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    program_id = request.GET.get('program_id') or request.GET.get('program')
    teacher_id = request.GET.get('teacher_id') or request.GET.get('teacher')
    room_id = request.GET.get('room_id') or request.GET.get('room')
    subject_id = request.GET.get('subject_id') or request.GET.get('subject')
    export_format = request.GET.get('format', 'json')

    if not start_date_str or not end_date_str:
        return Response({'error': 'start_date and end_date parameters are required'}, status=400)

    try:
        range_start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        range_end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)

    if range_end < range_start:
        return Response({'error': 'end_date must be on or after start_date'}, status=400)

    queryset = (
        Schedule.objects
        .filter(Q(week_start__lte=range_end) & Q(week_end__gte=range_start), is_active=True)
        .select_related('subject', 'teacher__user', 'room', 'program')
    )

    if program_id:
        queryset = queryset.filter(program_id=program_id)
    if teacher_id:
        queryset = queryset.filter(teacher_id=teacher_id)
    if room_id:
        queryset = queryset.filter(room_id=room_id)
    if subject_id:
        queryset = queryset.filter(subject_id=subject_id)

    user = request.user
    if user.role == 'department_head':
        queryset = queryset.filter(program__department=user.department)
    elif user.role == 'program_head':
        queryset = queryset.filter(program=user.program)
    elif user.role == 'teacher':
        queryset = queryset.filter(teacher__user=user)
    elif user.role == 'student':
        try:
            queryset = queryset.filter(program=user.student.program)
        except Exception:
            queryset = queryset.none()

    if export_format == 'pdf':
        return export_week_pdf(queryset, range_start, range_end)
    elif export_format == 'excel':
        return export_week_excel(queryset, range_start, range_end)

    days = []
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    total_sessions = 0
    total_minutes = 0

    current_date = range_start
    while current_date <= range_end:
        day_index = current_date.weekday()
        day_schedules = []
        for schedule in queryset.filter(day_of_week=day_index):
            total_sessions += 1
            duration_minutes = (
                datetime.combine(datetime.today(), schedule.end_time)
                - datetime.combine(datetime.today(), schedule.start_time)
            ).seconds // 60
            total_minutes += duration_minutes
            day_schedules.append({
                'id': schedule.id,
                'title': schedule.title,
                'subject_name': schedule.subject.name,
                'subject_code': getattr(schedule.subject, 'code', ''),
                'teacher_name': schedule.teacher.user.get_full_name(),
                'room_name': schedule.room.name,
                'room_capacity': schedule.room.capacity,
                'time_slot_info': {
                    'id': schedule.id,
                    'day_of_week': schedule.day_of_week,
                    'day_display': schedule.get_day_of_week_display(),
                    'start_time': schedule.start_time.strftime('%H:%M'),
                    'end_time': schedule.end_time.strftime('%H:%M'),
                    'duration_minutes': duration_minutes
                },
                'programs_list': [schedule.program.name],
                'start_date': current_date.strftime('%Y-%m-%d'),
                'end_date': current_date.strftime('%Y-%m-%d'),
                'duration_minutes': duration_minutes,
                'student_count': getattr(schedule.program, 'capacity', 0),
                'is_room_suitable': True,
                'is_cancelled': not schedule.is_active,
                'is_makeup': False,
                'notes': schedule.notes
            })
        day_schedules.sort(key=lambda x: x['time_slot_info']['start_time'])
        days.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'day_name': day_names[day_index],
            'schedules': day_schedules
        })
        current_date += timedelta(days=1)

    return Response({
        'start_date': range_start.strftime('%Y-%m-%d'),
        'end_date': range_end.strftime('%Y-%m-%d'),
        'days': days,
        'total_sessions': total_sessions,
        'total_hours': (total_minutes / 60.0) if total_minutes > 0 else 0
    })

# Absence Views
class AbsenceListCreateView(generics.ListCreateAPIView):
    serializer_class = AbsenceSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['absence_type', 'reason', 'is_approved']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return Absence.objects.all()
        elif user.role in ['department_head', 'program_head']:
            return Absence.objects.filter(
                Q(teacher__user__department=user.department) |
                Q(program__department=user.department)
            )
        elif user.role == 'teacher':
            return Absence.objects.filter(teacher__user=user)
        return Absence.objects.none()
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

class AbsenceDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = AbsenceSerializer
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return Absence.objects.all()
        elif user.role in ['department_head', 'program_head']:
            return Absence.objects.filter(
                Q(teacher__user__department=user.department) |
                Q(program__department=user.department)
            )
        elif user.role == 'teacher':
            return Absence.objects.filter(teacher__user=user)
        return Absence.objects.none()

# Makeup Session Views
class MakeupSessionListCreateView(generics.ListCreateAPIView):
    serializer_class = MakeupSessionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return MakeupSession.objects.all()
        elif user.role in ['department_head', 'program_head']:
            return MakeupSession.objects.filter(
                original_schedule__program__department=user.department
            )
        elif user.role == 'teacher':
            return MakeupSession.objects.filter(original_schedule__teacher__user=user)
        return MakeupSession.objects.none()
    
    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

class MakeupSessionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = MakeupSessionSerializer
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return MakeupSession.objects.all()
        elif user.role in ['department_head', 'program_head']:
            return MakeupSession.objects.filter(
                original_schedule__program__department=user.department
            )
        elif user.role == 'teacher':
            return MakeupSession.objects.filter(original_schedule__teacher__user=user)
        return MakeupSession.objects.none()

@api_view(['POST'])
def approve_makeup(request, pk):
    """Approve or reject makeup session"""
    try:
        makeup = MakeupSession.objects.get(pk=pk)
        action = request.data.get('action')  # 'approve' or 'reject'
        
        if action == 'approve':
            makeup.status = 'approved'
            makeup.approved_by = request.user
            message = 'Session de rattrapage approuvée'
        elif action == 'reject':
            makeup.status = 'rejected'
            makeup.approved_by = request.user
            message = 'Session de rattrapage rejetée'
        else:
            return Response({'error': 'Action invalide'}, status=400)
        
        makeup.save()
        
        return Response({
            'message': message,
            'makeup': MakeupSessionSerializer(makeup).data
        })
    
    except MakeupSession.DoesNotExist:
        return Response({'error': 'Session non trouvée'}, status=404)

@api_view(['GET'])
def available_rooms(request):
    """Get available rooms for a specific time slot"""
    day_of_week = request.GET.get('day_of_week')
    start_time = request.GET.get('start_time')
    end_time = request.GET.get('end_time')
    week_start = request.GET.get('week_start')
    week_end = request.GET.get('week_end')
    room_type = request.GET.get('room_type')
    
    if not all([day_of_week, start_time, end_time, week_start, week_end]):
        return Response({'error': 'Paramètres manquants'}, status=400)
    
    available = get_available_rooms(
        day_of_week=int(day_of_week),
        start_time=start_time,
        end_time=end_time,
        week_start=week_start,
        week_end=week_end,
        room_type=room_type
    )
    
    return Response(available)


@api_view(['GET'])
def get_schedule_by_week(request):
    """Get schedule for a specific week formatted for frontend"""
    # Support both week_start and start_date parameters for compatibility
    week_start = request.GET.get('week_start') or request.GET.get('start_date')
    program_id = request.GET.get('program_id') or request.GET.get('program')
    teacher_id = request.GET.get('teacher_id') or request.GET.get('teacher')
    room_id = request.GET.get('room_id') or request.GET.get('room')
    subject_id = request.GET.get('subject_id') or request.GET.get('subject')
    export_format = request.GET.get('format', 'json')
    
    if not week_start:
        return Response({'error': 'week_start or start_date parameter required'}, status=400)
    
    try:
        week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + timedelta(days=6)
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    
    # Base queryset - MODIFIÉ pour être plus flexible avec les dates
    queryset = Schedule.objects.filter(
        Q(week_start=week_start_date) | 
        Q(week_start__lte=week_start_date, week_end__gte=week_start_date),
        is_active=True
    ).select_related('subject', 'teacher__user', 'room', 'program')
    
    # Apply filters based on parameters
    if program_id:
        queryset = queryset.filter(program_id=program_id)
    if teacher_id:
        queryset = queryset.filter(teacher_id=teacher_id)
    if room_id:
        queryset = queryset.filter(room_id=room_id)
    if subject_id:
        queryset = queryset.filter(subject_id=subject_id)
    
    # Apply user role permissions
    user = request.user
    if user.role == 'department_head':
        queryset = queryset.filter(program__department=user.department)
    elif user.role == 'program_head':
        queryset = queryset.filter(program=user.program)
    elif user.role == 'teacher':
        queryset = queryset.filter(teacher__user=user)
    elif user.role == 'student':
        try:
            queryset = queryset.filter(program=user.student.program)
        except:
            queryset = queryset.none()
    
    # Handle different export formats FIRST
    if export_format == 'pdf':
        return export_week_pdf(queryset, week_start_date, week_end_date)
    elif export_format == 'excel':
        return export_week_excel(queryset, week_start_date, week_end_date)
    
    # Organize data by days for JSON response
    days_data = []
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    total_sessions = 0
    total_minutes = 0
    
    for day_index in range(7):
        current_date = week_start_date + timedelta(days=day_index)
        day_schedules = []
        
        # Get schedules for this day
        for schedule in queryset.filter(day_of_week=day_index):
            total_sessions += 1
            duration_minutes = (datetime.combine(datetime.today(), schedule.end_time) - 
                              datetime.combine(datetime.today(), schedule.start_time)).seconds // 60
            total_minutes += duration_minutes
            
            # Create schedule data structure matching frontend expectations
            schedule_data = {
                'id': schedule.id,
                'title': schedule.title,
                'subject_name': schedule.subject.name,
                'subject_code': getattr(schedule.subject, 'code', ''),
                'teacher_name': schedule.teacher.user.get_full_name(),
                'room_name': schedule.room.name,
                'room_capacity': schedule.room.capacity,
                'time_slot_info': {
                    'id': schedule.id,
                    'day_of_week': schedule.day_of_week,
                    'day_display': schedule.get_day_of_week_display(),
                    'start_time': schedule.start_time.strftime('%H:%M'),
                    'end_time': schedule.end_time.strftime('%H:%M'),
                    'duration_minutes': duration_minutes
                },
                'programs_list': [schedule.program.name],
                'start_date': current_date.strftime('%Y-%m-%d'),
                'end_date': current_date.strftime('%Y-%m-%d'),
                'duration_minutes': duration_minutes,
                'student_count': getattr(schedule.program, 'capacity', 0),
                'is_room_suitable': True,
                'is_cancelled': not schedule.is_active,
                'is_makeup': False,
                'notes': schedule.notes
            }
            day_schedules.append(schedule_data)
        
        # Sort schedules by start time
        day_schedules.sort(key=lambda x: x['time_slot_info']['start_time'])
        
        days_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'day_name': day_names[day_index],
            'schedules': day_schedules
        })
    
    # Calculate total hours
    total_hours = total_minutes / 60.0 if total_minutes > 0 else 0
    
    # Response data structure matching frontend expectations
    response_data = {
        'week_start': week_start_date.strftime('%Y-%m-%d'),
        'week_end': week_end_date.strftime('%Y-%m-%d'),
        'days': days_data,
        'total_sessions': total_sessions,
        'total_hours': total_hours
    }
    
    return Response(response_data)


def export_week_pdf(queryset, week_start_date, week_end_date):
    """Export weekly schedule to PDF"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from django.http import HttpResponse
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(f"Emploi du Temps - Semaine du {week_start_date} au {week_end_date}", styles['Title'])
        elements.append(title)
        
        # Prepare data for table
        data = [['Jour', 'Heure', 'Cours', 'Enseignant', 'Salle', 'Programme']]
        
        day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        
        for schedule in queryset.order_by('day_of_week', 'start_time'):
            day_name = day_names[schedule.day_of_week] if schedule.day_of_week < len(day_names) else f'Jour {schedule.day_of_week}'
            
            data.append([
                day_name,
                f"{schedule.start_time.strftime('%H:%M')}-{schedule.end_time.strftime('%H:%M')}",
                schedule.title[:30],  # Limit length
                schedule.teacher.user.full_name[:25],
                schedule.room.name[:15],
                schedule.program.name[:20]
            ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="emploi_temps_{week_start_date}_{week_end_date}.pdf"'
        
        return response
        
    except ImportError:
        return Response({'error': 'PDF export requires reportlab library'}, status=500)
    except Exception as e:
        return Response({'error': f'PDF export error: {str(e)}'}, status=500)


def export_week_excel(queryset, week_start_date, week_end_date):
    """Export weekly schedule to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Emploi du Temps"
        
        # Headers
        headers = ['Jour', 'Heure Début', 'Heure Fin', 'Cours', 'Enseignant', 'Salle', 'Programme']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        
        for row_num, schedule in enumerate(queryset.order_by('day_of_week', 'start_time'), 2):
            day_name = day_names[schedule.day_of_week] if schedule.day_of_week < len(day_names) else f'Jour {schedule.day_of_week}'
            
            data = [
                day_name,
                schedule.start_time.strftime('%H:%M'),
                schedule.end_time.strftime('%H:%M'),
                schedule.title,
                schedule.teacher.user.full_name,
                schedule.room.name,
                schedule.program.name
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="emploi_temps_{week_start_date}_{week_end_date}.xlsx"'
        
        return response
        
    except ImportError:
        return Response({'error': 'Excel export requires openpyxl library'}, status=500)
    except Exception as e:
        return Response({'error': f'Excel export error: {str(e)}'}, status=500)


def export_week_pdf(queryset, week_start_date, week_end_date):
    """Export weekly schedule to PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from django.http import HttpResponse
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Emploi du Temps - Semaine du {week_start_date} au {week_end_date}", styles['Title'])
    elements.append(title)
    
    # Prepare data for table
    data = [['Jour', 'Heure', 'Cours', 'Enseignant', 'Salle', 'Programme']]
    
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    for schedule in queryset.order_by('day_of_week', 'start_time'):
        day_name = day_names[schedule.day_of_week] if schedule.day_of_week < len(day_names) else f'Jour {schedule.day_of_week}'
        
        data.append([
            day_name,
            f"{schedule.start_time.strftime('%H:%M')}-{schedule.end_time.strftime('%H:%M')}",
            schedule.title[:30],  # Limit length
            schedule.teacher.user.full_name[:25],
            schedule.room.name[:15],
            schedule.program.name[:20]
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="emploi_temps_{week_start_date}_{week_end_date}.pdf"'
    
    return response


def export_week_excel(queryset, week_start_date, week_end_date):
    """Export weekly schedule to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    import io
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emploi du Temps"
    
    # Headers
    headers = ['Jour', 'Heure Début', 'Heure Fin', 'Cours', 'Enseignant', 'Salle', 'Programme']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    day_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
    
    for row_num, schedule in enumerate(queryset.order_by('day_of_week', 'start_time'), 2):
        day_name = day_names[schedule.day_of_week] if schedule.day_of_week < len(day_names) else f'Jour {schedule.day_of_week}'
        
        data = [
            day_name,
            schedule.start_time.strftime('%H:%M'),
            schedule.end_time.strftime('%H:%M'),
            schedule.title,
            schedule.teacher.user.full_name,
            schedule.room.name,
            schedule.program.name
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="emploi_temps_{week_start_date}_{week_end_date}.xlsx"'
    
    return response
